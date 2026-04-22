from flask import Flask, request, send_file
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import pandas as pd

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route("/")
def index():
    return '''
    <h2>Excel 合併工具（優化版）</h2>
    <form method="post" action="/merge" enctype="multipart/form-data">
        <input type="file" name="files" multiple>
        <br><br>
        <button type="submit">上傳並合併</button>
    </form>
    '''

@app.route("/merge", methods=["POST"])
def merge():
    files = request.files.getlist("files")

    # 櫃號封條優先排序
    files = sorted(files, key=lambda x: ("櫃號封條" not in x.filename, x.filename))

    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    for file in files:
        filename = secure_filename(file.filename)

        if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
            continue

        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        try:
            # ===== 處理 .xls =====
            if filename.endswith(".xls"):
    import xlrd
    book = xlrd.open_workbook(filepath)

    # 櫃號封條 → 只取第一個sheet
    if "櫃號封條" in filename:
        sheet_list = [book.sheet_by_index(0)]
    else:
        sheet_list = [book.sheet_by_index(i) for i in range(book.nsheets)]

    for sheet in sheet_list:
        new_sheet = new_wb.create_sheet(title=sheet.name[:31])

        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                new_sheet.cell(row=r+1, column=c+1, value=sheet.cell_value(r, c))

            # ===== 處理 .xlsx =====
            else:
                wb = load_workbook(filepath)

            # ===== 櫃號封條邏輯 =====
            if "櫃號封條" in filename:
                sheets = [wb.worksheets[0]]
            else:
                sheets = wb.worksheets

            # ===== 複製內容 =====
            for sheet in sheets:
                new_sheet = new_wb.create_sheet(title=sheet.title[:31])

                for row in sheet.iter_rows():
                    for cell in row:
                        new_sheet[cell.coordinate].value = cell.value

                # 自動調整欄寬（優化排版）
                for col in new_sheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    new_sheet.column_dimensions[col_letter].width = max_length + 2

        except Exception as e:
            print("錯誤:", e)
            continue

    output_path = os.path.join(UPLOAD_FOLDER, "merged.xlsx")
    new_wb.save(output_path)

    return send_file(output_path, as_attachment=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
